import os
import sqlite3
from tkinter import *
#import xlsxwriter
from datetime import datetime, timedelta
import cv2
import time
from openpyxl import load_workbook
from PIL import ImageTk, Image
conn = sqlite3.connect('form1.db')
c = conn.cursor()


#next level code for excel sheet updation
filename="hello.xlsx"
wb=load_workbook(filename)
ws=wb.worksheets[0]

i=2
c.execute('SELECT * FROM attendance_sheet')
result_set = c.fetchall()
conn.commit()

for row in result_set:
   i=i+1
   ws["A"+str(i)]=row[0]
   print("hii...")
   ws["B"+str(i)]=row[1]
   ws["C"+str(i)]=row[2]
   ws["D"+str(i)]=row[3]
   ws["E"+str(i)]=row[4]
   ws["F"+str(i)]=row[5]
   ws["G"+str(i)]=row[6]
   if(row[7]):
      ws["H"+str(i)]=row[7]
   wb.save(filename)
   conn.close()
   

