import os
import sqlite3
from tkinter import *
#import xlsxwriter
from datetime import datetime, timedelta
import cv2
import time
from openpyxl import load_workbook
from PIL import ImageTk, Image
conn = sqlite3.connect('form1.db')
c = conn.cursor()
filepath = "recognizer/trainingData.yml"
first_id="xyz"
fid="pqr"
ottm='123'
intime='123'

def popupmsg(msg):
   popup=Tk()
   popup.wm_title("!")
   popup.resizable(0,0)
   label=Label(popup,text=msg,width=40,font=("bold", 10))
   label.pack(side="top",fill="x",pady=10)
   B1=Button(popup,text="OK",command=popup.destroy)
   B1.pack()
   popup.mainloop()

#url="http://192.168.43.1:8080/shot.jpg"  #android part

'''if not os.path.isfile(filepath):
  print("Please train the data first")
  exit(0)'''
face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(filepath)
cap = cv2.VideoCapture(0,cv2.CAP_DSHOW)  #for capturing contents from live video
#cap=cv2.VideoCapture('vdi.mp4',cv2.CAP_DSHOW)   #for capturing contents from recorded video
test = "no"
count = 1
#while(cap.isOpened()):   #for capturing contents from recorded video
#while True:              #for capturing contents from live video
while True:
   if(test == "yes"):
      break
   # datetime object containing current date and time
   now = datetime.now()
   current_date = now.strftime("%d/%m/%Y")
   current_time = now.strftime("%H:%M:%S")
   #img_resp=requests.get(url)#android part
   ret, img = cap.read()
   cv2.putText(img, str("Please Look into the camera"), (100, 25), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255),2)
   #img_arr=np.array(bytearray(img_resp.content),dtype=np.uint8) #android part
   #img=cv2.imdecode(img_arr, -1)  #android part
   gray = cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
   faces = face_cascade.detectMultiScale(gray, 1.3, 5)
   for (x,y,w,h) in faces:
      cv2.rectangle(img,(x,y),(x+w,y+h),(255,255,0),2) # x,y top left corner  x+w=left+width  y+h=top+height  255,255,0=color  2=thickness of rectangle border
      ids,conf = recognizer.predict(gray[y:y+h,x:x+w])
      if(fid == "pqr"):
         fid=ids
         if conf < 50:
            c.execute('SELECT * FROM student WHERE id=?', (ids,))
            result = c.fetchall()
            conn.commit()
            for row in result:
               id = row[0]
               fname = row[1]
               mname = row[2]
               lname = row[3]
               dept = row[8]
               cv2.putText(img, fname, (x + 5, y+h-5), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255, 0), 2)
               rowno = 0
               c.execute('SELECT * FROM attendance_sheet WHERE id=? AND date=? ORDER BY datetime(date) DESC Limit 1', (id,current_date,))
               result_set = c.fetchall()
               for row in result_set:
                  rowno = rowno + 1
                  if(row[7]):
                     ottm = row[7]
                  if(row[6]):
                     intime = row[6]
               conn.commit()
               if (rowno > 0):
                  print("old time:",ottm)
                  #if outtime is updated priviously then bellow code will execute otherwise else part will execute
                  oldtime1 = datetime.now() - timedelta(minutes=15)
                  time1 = oldtime1.strftime("%H:%M:%S")#time before 15 minutes than current time
                  print("Before 15 minute:",time1)
                  oldtime2 = datetime.now() - timedelta(minutes=3)
                  time2 = oldtime2.strftime("%H:%M:%S")#time before 3 minutes than current time
                  print("Before 3 minute:",time2)
                  if(ottm=='123' and intime<time1):
                     c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=?', (current_time, current_date, id))
                     print("///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////")
                     #above code if outtime is not present then current time will be updated as outtime
                     conn.commit()
                     popupmsg("Attendance Out Update Success!")
                     time.sleep(70)
                  elif(ottm=='0' and intime<time1):
                     c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=?', (current_time, current_date, id))
                     print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                     print("~~~~~~~ In Time:",intime)
                     print("~~~~~~~ Before 3 minute:",time1)
                     #this code will execute if outtimeis 0 then current time will be updated as outtime
                     conn.commit()
                     popupmsg("Attendance Out Update Success!")
                     time.sleep(70)
                  elif(ottm>time1 and intime<time1):
                     if(ottm<time2):
                        c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=? AND outtime>?  AND outtime<?', (0, current_date, id, time1, time2))
                        print("*****************************************************************************************************************************************")
                        #above code if outtime is greater than time before 15 minute then only outtime will be updated
                        conn.commit()
                        popupmsg("Attendance Out Update Success!")
                        time.sleep(70)
                  else:
                     oldtime = datetime.now() - timedelta(minutes=2)
                     #here above code will extract time before two minute and it will store inside oldtime variable
                     #this code is execute when he will not return withing 15 minutes
		     #time = oldtime.strftime("%H:%M:%S")
                     #c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=? AND intime<?', (current_time, current_date, id, time,))
                     print("000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000")
                     #above code if intime is less than current time by two 2 minute then only outtime will be updated
                     #conn.commit()
                     popupmsg("Sorry You are Late!")
                     print(fname)
                     print("Out time")
                     print(current_time)
                     time.sleep(70)
               else:
                  c.execute('INSERT INTO attendance_sheet (id,fname,mname,lname,dept,intime,date) VALUES(?,?,?,?,?,?,?)',(id, fname, mname, lname, dept, current_time, current_date,))
                  conn.commit()
                  print(fname)
                  print("In time")
                  print(current_time)
                  popupmsg("Attendance In Update Success!")
                  time.sleep(70)
         else:
            cv2.putText(img,'No Match', (x+2,y+h-5), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255),2)
      elif(fid!="pqr" and fid == ids ):
         fid=ids
         if conf < 50:
            c.execute('SELECT * FROM student WHERE id=?', (ids,))
            result = c.fetchall()
            conn.commit()
            for row in result:
               id = row[0]
               fname = row[1]
               mname = row[2]
               lname = row[3]
               dept = row[8]
               cv2.putText(img, fname, (x + 5, y+h-5), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255, 0), 2)
               rowno = 0
               c.execute('SELECT * FROM attendance_sheet WHERE id=? AND date=? ORDER BY datetime(date) DESC Limit 1', (id,current_date,))
               result_set = c.fetchall()
               for row in result_set:
                  rowno = rowno + 1
                  if(row[7]):
                     ottm = row[7]
                     print("This is ottm::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::")
                     print(ottm)
                  if(row[6]):
                     intime = row[6]
               conn.commit()
               if (rowno > 0):
                  #if outtime is updated priviously then bellow code will execute otherwise else part will execute
                  print("old time:",ottm)
                  oldtime1 = datetime.now() - timedelta(minutes=15)
                  time1 = oldtime1.strftime("%H:%M:%S")#time before 15 minutes than current time
                  print("Before 15 minute:",time1)
                  oldtime2 = datetime.now() - timedelta(minutes=3)
                  time2 = oldtime2.strftime("%H:%M:%S")#time before 3 minutes than current time
                  print("Before 3 minute:",time2)
                  if(ottm=='123' and intime<time1):
                     c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=?', (current_time, current_date, id))
                     print("///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////")
                     print("////// In Time:",intime)
                     print("////// Bifore 3 minute:",time1)
                     #above code if outtime is not present then current time will be updated as outtime
                     conn.commit()
                     popupmsg("Attendance Out Update Success!")
                     time.sleep(70)
                  elif(ottm=='0' and intime<time1):
                     c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=?', (current_time, current_date, id))
                     print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                     print("~~~~~~~ In Time:",intime)
                     print("~~~~~~~ Before 3 minute:",time1)
                     #this code will execute if outtimeis 0 then current time will be updated as outtime
                     conn.commit()
                     popupmsg("Attendance Out Update Success!")
                     time.sleep(70)
                  elif(ottm>time1 and intime<time1):
                     if(ottm<time2):
                        c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=? AND outtime>?  AND outtime<?', (0, current_date, id, time1, time2))
                        #c.execute('UPDATE attendance_sheet SET outtime=? WHERE date=? AND id=? AND outtime>?  AND outtime<?', (current_time, current_date, id, time1, time2))
                        print("***********************************************deleting outtime******************************************************************************************")
                        #above code if outtime is greater than time before 15 minute means student returns before 15 minutes
                        #then only privious set outtime will be deleted
                        conn.commit()
                        popupmsg("Attendance Out Cancel Success!")
                        time.sleep(70)
                  else:
                     oldtime = datetime.now() - timedelta(minutes=2)
                     #here above code will extract time before two minute and it will store inside oldtime variable
                     #this code will execute when he will not return withing 15 minutes
                     print("000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000")
                     #above code if intime is less than current time by two 2 minute then only outtime will be updated
                     #conn.commit()
                     popupmsg("Sorry You are Late!")
                     print(fname)
                     print("Out time")
                     print(current_time)
                     time.sleep(70)
               else:
                  c.execute('INSERT INTO attendance_sheet (id,fname,mname,lname,dept,intime,date) VALUES(?,?,?,?,?,?,?)',(id, fname, mname, lname, dept, current_time, current_date,))
                  conn.commit()
                  print(fname)
                  print("In time")
                  print(current_time)
                  popupmsg("Attendance In Update Success!")
                  time.sleep(70)
         else:
            cv2.putText(img,'No Match', (x+2,y+h-5), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255),2)
      elif(fid != "pqr" and fid != ids and count < 15 ):
         count = count + 1
         print("########################################################")
         print(count)
      elif(fid != "pqr" and fid != ids and count > 14 ):
         print("iiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiii")
         print(ids)
         print("********************************************************************")
         test = "yes"
         
         break
         cv2.destroyAllWindows()
         
         
         
         
      
      
   cv2.imshow('Face Recognizer',img)
   k = cv2.waitKey(30) & 0xff
   if k == 27:
      break
#cap.release()
cv2.destroyAllWindows()

if(test != "yes"):
   #next level code for excel sheet updation
   filename="hello.xlsx"
   wb=load_workbook(filename)
   ws=wb.worksheets[0]

   i=2
   c.execute('SELECT * FROM attendance_sheet')
   result_set = c.fetchall()
   conn.commit()

   for row in result_set:
    i=i+1
    ws["A"+str(i)]=row[0]
    ws["B"+str(i)]=row[1]
    ws["C"+str(i)]=row[2]
    ws["D"+str(i)]=row[3]
    ws["E"+str(i)]=row[4]
    ws["F"+str(i)]=row[5]
    ws["G"+str(i)]=row[6]
    if(row[7]):
      ws["H"+str(i)]=row[7]
   wb.save(filename)
   conn.close()
   os.system('home.py')
   

